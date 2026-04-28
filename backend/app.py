"""
GLOF Early Warning System — Flask Application (Production-Grade)
Features:
  - Input validation (marshmallow)
  - Rate limiting (flask-limiter)
  - HMAC-SHA256 sensor authentication
  - Structured JSON logging
  - Prometheus metrics
  - Swagger / OpenAPI docs
  - Alert state-machine (OPEN → ACKNOWLEDGED → RESOLVED)
  - Audit logging
  - Trend analysis + data export endpoints
  - Improved /health endpoint
"""
import os
import sys

# ── Ensure this file's directory is in sys.path so local packages
# (core, routes, models, tasks) resolve correctly under gunicorn. ──────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)
import json
import time
import hmac
import hashlib
import csv
import io
from datetime import datetime, timedelta, timezone

from flask import Flask, Response, request, jsonify, stream_with_context
from flask_cors import CORS
from flask_jwt_extended import JWTManager, jwt_required, get_jwt_identity, get_jwt
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flasgger import Swagger
from prometheus_flask_exporter import PrometheusMetrics
from prometheus_client import Counter, Gauge
from pymongo import MongoClient, DESCENDING
from bson import ObjectId
import redis
from dotenv import load_dotenv

from routes.auth import auth_bp, init_auth
from routes.data import lakes_bp, events_bp, alerts_bp, init_lakes, init_events, init_alerts
from models.seed import seed_database
from core.risk_engine import calculate_risk, should_alert
from core.schemas import TelemetrySchema, validate_json
from core.logger import telemetry_log, alert_log, audit_log, get_logger
from core.middleware import register_security_headers, register_error_handlers
from core.db_indexes import ensure_indexes
try:
    from tasks import enqueue_email_jobs, enqueue_alert_emails, enqueue_alert_push_notifications, enqueue_alert_sms
    _CELERY_AVAILABLE = True
except ImportError:
    _CELERY_AVAILABLE = False
    def enqueue_email_jobs(*args, **kwargs):  # noqa: E301
        """Fallback when Celery is unavailable."""
        return {"queued": 0, "skipped": []}
    def enqueue_alert_emails(*args, **kwargs):  # noqa: E301
        """Fallback when Celery is unavailable."""
        return {"queued": 0, "skipped": []}
    def enqueue_alert_push_notifications(*args, **kwargs):  # noqa: E301
        """Fallback when Celery is unavailable."""
        return {"queued": 0}
    def enqueue_alert_sms(*args, **kwargs):  # noqa: E301
        """Fallback when Celery is unavailable."""
        return {"queued": 0, "skipped": []}

load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
load_dotenv()

app_log = get_logger("glof.app")

# ─── App Setup ────────────────────────────────────────────────────────────────
app = Flask(__name__)

# Parse CORS origins — env var may be a comma-separated string or a single URL
_raw_origins = os.environ.get("CORS_ORIGINS", "*")
_cors_origins = [o.strip() for o in _raw_origins.split(",")] if "," in _raw_origins else _raw_origins

CORS(
    app,
    resources={r"/api/*": {"origins": _cors_origins}},
    supports_credentials=True,
    allow_headers=["Content-Type", "Authorization", "X-Sensor-Signature"],
    methods=["GET", "POST", "PATCH", "PUT", "DELETE", "OPTIONS"],
    expose_headers=["Content-Type", "Authorization"],
)

# Swagger / OpenAPI
app.config["SWAGGER"] = {
    "title": "GLOF Early Warning System API",
    "uiversion": 3,
    "description": "Real-time Glacial Lake Outburst Flood monitoring API",
    "version": "2.0.0",
}
swagger = Swagger(app)

# JWT — fail on startup in production if secret is not set
_jwt_secret = os.environ.get("JWT_SECRET_KEY", "change-me-in-prod")
if os.environ.get("FLASK_ENV") == "production" and _jwt_secret == "change-me-in-prod":
    raise RuntimeError("JWT_SECRET_KEY must be set in production!")
app.config["JWT_SECRET_KEY"] = _jwt_secret
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(hours=24)
jwt = JWTManager(app)

# Rate Limiter — uses REDIS_URL (works for both Upstash rediss:// and local redis://)
_redis_url = os.environ.get("REDIS_URL", "redis://localhost:6379/0")

limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["200 per minute"],
    storage_uri="memory://",
)

# Prometheus Metrics
metrics = PrometheusMetrics(app, group_by="endpoint")
TELEMETRY_COUNTER = Counter(
    "glof_telemetry_received_total", "Total telemetry readings received"
)
ALERT_COUNTER = Counter(
    "glof_alerts_fired_total", "Total alerts fired", ["level"]
)
RISK_GAUGE = Gauge(
    "glof_lake_risk_score", "Current risk score per lake", ["lake_id"]
)

# ─── Database ─────────────────────────────────────────────────────────────────
MONGO_URI = os.environ.get("MONGO_URI", "mongodb://localhost:27017/glof_db")
mongo_client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
db = mongo_client.glof_db

redis_client = redis.from_url(_redis_url, decode_responses=True)

# Sensor HMAC key (IoT authentication)
SENSOR_API_KEY = os.environ.get("SENSOR_API_KEY", "glof-sensor-secret-key-2024")
ALERT_COOLDOWN_SECONDS = int(os.environ.get("ALERT_COOLDOWN_SECONDS", 900))


# ─── Helpers ──────────────────────────────────────────────────────────────────
def verify_sensor_hmac(request_data: dict, signature: str) -> bool:
    """Verify HMAC-SHA256 signature from IoT sensor."""
    expected = hmac.new(
        SENSOR_API_KEY.encode(),
        json.dumps(request_data, sort_keys=True).encode(),
        hashlib.sha256,
    ).hexdigest()
    return hmac.compare_digest(expected, signature)


def write_audit_log(action: str, actor: str, details: dict = None):
    """Write an audit entry to MongoDB and the audit logger."""
    entry = {
        "action": action,
        "actor": actor,
        "details": details or {},
        "timestamp": datetime.now(tz=timezone.utc).isoformat(),
        "remote_addr": request.remote_addr,
    }
    db.audit_logs.insert_one(entry)
    audit_log.info(action, extra={"user": actor, **{k: v for k, v in (details or {}).items() if isinstance(v, (str, int, float))}})


def is_alert_suppressed(lake_id: str, alert_type: str) -> bool:
    """Suppress duplicate alerts for the same lake/type inside the cooldown window."""
    cooldown_key = f"alert_cooldown:{lake_id}:{alert_type.lower()}"
    try:
        if redis_client.get(cooldown_key):
            return True
        redis_client.setex(cooldown_key, ALERT_COOLDOWN_SECONDS, "1")
        return False
    except Exception as e:
        app_log.warning(f"Redis cooldown error, degrading gracefully: {e}")
        return False


def create_demo_users(database):
    """Create default demo accounts if they don't exist."""
    import bcrypt

    defaults = [
        {"email": "admin@glof.in", "password": "admin123", "name": "NDMA Admin", "role": "admin"},
        {"email": "user@glof.in", "password": "user123", "name": "Researcher", "role": "user"},
    ]

    # Allow overriding via env vars so the real admin email/password can be set
    # in the Render dashboard without touching code.
    _env_admin_email = os.environ.get("ADMIN_EMAIL", "").strip().lower()
    _env_admin_password = os.environ.get("ADMIN_PASSWORD", "").strip()
    if _env_admin_email and _env_admin_password:
        defaults.append({
            "email": _env_admin_email,
            "password": _env_admin_password,
            "name": os.environ.get("ADMIN_NAME", "Admin"),
            "role": "admin",
        })

    for u in defaults:
        if not database.users.find_one({"email": u["email"]}):
            hashed = bcrypt.hashpw(u["password"].encode(), bcrypt.gensalt())
            database.users.insert_one({
                "email": u["email"], "name": u["name"], "role": u["role"],
                "password": hashed, "created_at": datetime.now(tz=timezone.utc),
                "alert_preferences": {
                    "warnings_enabled": True,
                    "emergencies_enabled": True,
                    "email_enabled": True,
                },
            })
            app_log.info(f"Created demo user: {u['email']}")


# ─── Startup (non-blocking) ──────────────────────────────────────────────────
import threading

def _run_startup_tasks():
    try:
        ensure_indexes(db)
    except Exception as _e:
        app_log.warning(f"ensure_indexes skipped: {_e}")

    try:
        seed_database(db)
    except Exception as _e:
        app_log.warning(f"seed_database skipped: {_e}")

    try:
        create_demo_users(db)
    except Exception as _e:
        app_log.warning(f"create_demo_users skipped: {_e}")

# Start initialization in background so Gunicorn healthchecks don't timeout
threading.Thread(target=_run_startup_tasks, daemon=True).start()

register_security_headers(app)
register_error_handlers(app)

# Register blueprints
app.register_blueprint(init_auth(db), url_prefix="/api/auth")
app.register_blueprint(init_lakes(db), url_prefix="/api/lakes")
app.register_blueprint(init_events(db), url_prefix="/api/events")
app.register_blueprint(init_alerts(db, redis_client), url_prefix="/api/alerts")


@app.route("/")
def root():
    """Root health-check — confirms the GLOF Flask auth app is running."""
    return jsonify({"service": "GLOF Early Warning System API", "version": "2.1.0", "status": "ok"})



# ─── Telemetry Endpoint ────────────────────────────────────────────────────────
@app.route("/api/telemetry", methods=["POST"])
@limiter.limit("500 per minute")
def receive_telemetry():
    """
    Ingest sensor telemetry reading.
    ---
    tags:
      - Telemetry
    parameters:
      - in: body
        name: body
        required: true
        schema:
          properties:
            lake_id: {type: string, example: GL001}
            lake_name: {type: string}
            temperature: {type: number, example: 12.5}
            rainfall: {type: number, example: 45.2}
            water_level_rise: {type: number, example: 120.0}
    responses:
      200:
        description: Risk score and alert status
      400:
        description: Validation error
      401:
        description: Invalid sensor signature
    """
    raw_data = request.get_json()
    if not raw_data:
        return jsonify({"error": "No JSON body"}), 400

    # Optional HMAC verification (enabled only if header present — backward compat)
    sig = request.headers.get("X-Sensor-Signature")
    if sig and not verify_sensor_hmac(raw_data, sig):
        app_log.warning("Invalid sensor signature rejected", extra={"remote_addr": request.remote_addr})
        return jsonify({"error": "Invalid sensor signature"}), 401

    # Validate inputs
    data, errors = validate_json(TelemetrySchema, raw_data)
    if errors:
        return jsonify({"error": "Validation failed", "details": errors}), 422

    lake_id = data["lake_id"]
    temp = data["temperature"]
    rainfall = data["rainfall"]
    wl_rise = data["water_level_rise"]
    timestamp = data.get("timestamp") or datetime.now(tz=timezone.utc).isoformat()

    # ── Velocity feature (rate-of-change from Redis rolling window) ──────────
    window_key = f"wl_window:{lake_id}"
    try:
        redis_client.rpush(window_key, wl_rise)
        redis_client.ltrim(window_key, -10, -1)  # Keep last 10 readings
        redis_client.expire(window_key, 3600)
        wl_history = [float(v) for v in redis_client.lrange(window_key, 0, -1)]
    except Exception as e:
        app_log.warning(f"Redis velocity error: {e}")
        wl_history = [wl_rise]

    velocity = 0.0
    if len(wl_history) >= 2:
        velocity = (wl_history[-1] - wl_history[0]) / max(len(wl_history) - 1, 1)

    # ── Risk Calculation ──────────────────────────────────────────────────────
    risk = calculate_risk(temp, rainfall, wl_rise, velocity=velocity)
    alert_info = should_alert(risk)

    # ── Build telemetry document ──────────────────────────────────────────────
    tel_doc = {
        "lake_id": lake_id,
        "lake_name": data["lake_name"],
        "timestamp": timestamp,
        "temperature": temp,
        "rainfall": rainfall,
        "water_level_rise": wl_rise,
        "velocity": round(velocity, 3),
        "risk_score": risk["score"],
        "risk_level": risk["level"],
        "risk_color": risk["color"],
        "breakdown": risk["breakdown"],
        "ml_score": risk.get("ml_score"),
    }

    # ── Persist ───────────────────────────────────────────────────────────────
    db.telemetry.insert_one({**tel_doc})
    db.lakes.update_one({"id": lake_id}, {"$set": {
        "current_risk_score": risk["score"],
        "current_risk_level": risk["level"],
        "last_updated": datetime.now(tz=timezone.utc).isoformat(),
    }})

    # ── Prometheus ────────────────────────────────────────────────────────────
    TELEMETRY_COUNTER.inc()
    RISK_GAUGE.labels(lake_id=lake_id).set(risk["score"])

    # ── Alert ─────────────────────────────────────────────────────────────────
    if alert_info["alert"]:
        alert_suppressed = is_alert_suppressed(lake_id, alert_info["type"])
        alert_doc = {
            "lake_id": lake_id,
            "lake_name": data["lake_name"],
            "type": alert_info["type"],
            "message": alert_info["message"],
            "risk_score": risk["score"],
            "risk_level": risk["level"],
            "timestamp": timestamp,
            "status": "OPEN",            # State-machine field
            "resolved": False,
            "is_test": False,
            "cooldown_suppressed": alert_suppressed,
        }
        if not alert_suppressed:
            db.alerts.insert_one(alert_doc)
            ALERT_COUNTER.labels(level=risk["level"]).inc()
            alert_log.warning(
                f"Alert fired: {alert_info['type']}",
                extra={"lake_id": lake_id, "risk_score": risk["score"], "alert_type": alert_info["type"]},
            )

            # ── Email all opted-in users for Warning (≥61) and Emergency (≥80) ──
            # Fetch the lake record so we can add state/elevation to the email
            lake_doc = db.lakes.find_one({"id": lake_id}, {"_id": 0,
                "state": 1, "elevation_m": 1}) or {}

            alert_payload = {
                "lake_id":       lake_id,
                "lake_name":     data["lake_name"],
                "risk_level":    risk["level"],
                "risk_score":    risk["score"],
                "alert_type":    alert_info["type"],
                "alert_message": alert_info["message"],
                "breakdown":     risk.get("breakdown", {}),
                "timestamp":     timestamp,
                "lake_state":    lake_doc.get("state", ""),
                "lake_elevation": str(lake_doc.get("elevation_m", "")),
            }

            # Emergency alerts → all opted-in users
            # Warning alerts   → opted-in users only
            users = list(db.users.find(
                {"alert_preferences.email_enabled": True, "email": {"$ne": ""}},
                {"email": 1, "phone": 1, "_id": 0},
            ))

            if users:
                q = enqueue_alert_emails(users, alert_payload)
                push_q = enqueue_alert_push_notifications(alert_payload)
                
                # Send SMS for Critical and Emergency alerts
                sms_q = {"queued": 0, "skipped": []}
                if risk["level"] in ["Critical", "Emergency"]:
                    sms_q = enqueue_alert_sms(users, alert_payload)
                
                app_log.info(
                    "Alert notifications queued",
                    extra={
                        "lake_id":    lake_id,
                        "alert_type": alert_info["type"],
                        "risk_score": risk["score"],
                        "emails_queued": q["queued"],
                        "push_queued": push_q["queued"],
                        "sms_queued": sms_q["queued"]
                    },
                )

        else:
            alert_info = {**alert_info, "suppressed": True, "cooldown_seconds": ALERT_COOLDOWN_SECONDS}

    # ── Redis pub/sub ─────────────────────────────────────────────────────────
    payload = {**tel_doc, "alert": alert_info}
    try:
        redis_client.publish("glof_stream", json.dumps(payload, default=str))
    except Exception as e:
        app_log.warning(f"Redis publish error: {e}")

    telemetry_log.info(
        f"Telemetry received: {lake_id}",
        extra={"lake_id": lake_id, "risk_score": risk["score"]},
    )
    return jsonify({"status": "ok", "risk": risk, "alert": alert_info}), 200


# ─── SSE Stream ───────────────────────────────────────────────────────────────
@app.route("/api/stream")
def sse_stream():
    """
    Server-Sent Events live data stream.
    ---
    tags:
      - Streaming
    responses:
      200:
        description: text/event-stream of telemetry + alert events
    """
    def event_generator():
        try:
            pubsub = redis_client.pubsub()
            pubsub.subscribe("glof_stream")
            yield 'data: {"type": "connected"}\n\n'
            for message in pubsub.listen():
                if message["type"] == "message":
                    yield f"data: {message['data']}\n\n"
        except Exception as e:
            app_log.warning(f"SSE Redis error: {e}")
            yield 'data: {"type": "connected"}\n\n'
            while True:
                time.sleep(15)
                yield 'data: {"type": "ping"}\n\n'

    # Determine the allowed origin for this response
    req_origin = request.headers.get("Origin", "")
    allowed_origins = [o.strip() for o in os.environ.get("CORS_ORIGINS", "*").split(",")]
    if "*" in allowed_origins:
        sse_origin = "*"
    elif req_origin in allowed_origins:
        sse_origin = req_origin
    else:
        sse_origin = allowed_origins[0] if allowed_origins else "*"

    return Response(
        stream_with_context(event_generator()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Access-Control-Allow-Origin": sse_origin,
            "Access-Control-Allow-Credentials": "true",
        },
    )


# ─── Dashboard Summary ────────────────────────────────────────────────────────
@app.route("/api/dashboard/summary", methods=["GET"])
@jwt_required()
def dashboard_summary():
    """
    Aggregated summary for dashboard cards.
    ---
    tags:
      - Dashboard
    security:
      - Bearer: []
    responses:
      200:
        description: Summary stats
    """
    total_lakes = db.lakes.count_documents({})
    critical = db.lakes.count_documents({"current_risk_level": "Critical"})
    high = db.lakes.count_documents({"current_risk_level": "High"})
    active_alerts = db.alerts.count_documents({"resolved": False})
    total_events = db.glof_events.count_documents({})

    latest_telemetry = list(db.telemetry.find({}, {"_id": 0}).sort("timestamp", DESCENDING).limit(1))

    return jsonify({
        "total_lakes": total_lakes,
        "critical_lakes": critical,
        "high_risk_lakes": high,
        "active_alerts": active_alerts,
        "total_events": total_events,
        "latest_reading": latest_telemetry[0] if latest_telemetry else None,
    }), 200


# ─── Trend Analysis ───────────────────────────────────────────────────────────
@app.route("/api/lakes/<lake_id>/trend", methods=["GET"])
@jwt_required()
def lake_trend(lake_id):
    """
    Returns 6-hour trend and velocity for a lake.
    ---
    tags:
      - Lakes
    parameters:
      - name: lake_id
        in: path
        type: string
        required: true
    security:
      - Bearer: []
    responses:
      200:
        description: Trend statistics
    """
    since = datetime.now(tz=timezone.utc) - timedelta(hours=6)
    readings = list(
        db.telemetry.find(
            {"lake_id": lake_id, "timestamp": {"$gte": since.isoformat()}},
            {"_id": 0, "water_level_rise": 1, "risk_score": 1, "timestamp": 1},
        ).sort("timestamp", 1)
    )

    if len(readings) < 2:
        return jsonify({
            "lake_id": lake_id,
            "trend": "insufficient_data",
            "velocity_cm_per_reading": 0,
            "risk_delta": 0,
            "readings_count": len(readings),
        }), 200

    first_wl = readings[0]["water_level_rise"]
    last_wl = readings[-1]["water_level_rise"]
    first_risk = readings[0]["risk_score"]
    last_risk = readings[-1]["risk_score"]

    velocity = (last_wl - first_wl) / max(len(readings) - 1, 1)
    risk_delta = last_risk - first_risk

    if velocity > 2:
        trend = "rising_fast"
    elif velocity > 0.5:
        trend = "rising"
    elif velocity < -0.5:
        trend = "falling"
    else:
        trend = "stable"

    return jsonify({
        "lake_id": lake_id,
        "trend": trend,
        "velocity_cm_per_reading": round(velocity, 3),
        "risk_delta": round(risk_delta, 2),
        "readings_count": len(readings),
        "first_wl": first_wl,
        "last_wl": last_wl,
        "first_risk": first_risk,
        "last_risk": last_risk,
    }), 200


# ─── Data Export ──────────────────────────────────────────────────────────────
@app.route("/api/lakes/<lake_id>/export", methods=["GET"])
@jwt_required()
def export_telemetry(lake_id):
    """
    Export telemetry data as CSV or JSON.
    ---
    tags:
      - Lakes
    parameters:
      - name: lake_id
        in: path
        type: string
        required: true
      - name: format
        in: query
        type: string
        enum: [csv, json]
        default: json
      - name: from
        in: query
        type: string
        description: ISO timestamp
      - name: to
        in: query
        type: string
        description: ISO timestamp
    security:
      - Bearer: []
    responses:
      200:
        description: Data file
    """
    fmt = request.args.get("format", "json").lower()
    from_ts = request.args.get("from")
    to_ts = request.args.get("to")

    query = {"lake_id": lake_id}
    if from_ts or to_ts:
        query["timestamp"] = {}
        if from_ts:
            query["timestamp"]["$gte"] = from_ts
        if to_ts:
            query["timestamp"]["$lte"] = to_ts

    rows = list(db.telemetry.find(query, {"_id": 0}).sort("timestamp", 1))

    if fmt == "csv":
        if not rows:
            return Response("No data", mimetype="text/plain"), 204

        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=["lake_id", "lake_name", "timestamp", "temperature",
                        "rainfall", "water_level_rise", "velocity",
                        "risk_score", "risk_level"],
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(rows)
        csv_content = output.getvalue()

        return Response(
            csv_content,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={lake_id}_telemetry.csv"},
        )

    return jsonify(rows), 200


# ─── Bulk Telemetry Export (all lakes, time window) ───────────────────────────
@app.route("/api/telemetry/export", methods=["GET"])
@jwt_required()
def export_bulk_telemetry():
    """
    Export last N hours of telemetry for all lakes (or a specific lake) as CSV or XLSX.
    ---
    tags:
      - Telemetry
    parameters:
      - name: hours
        in: query
        type: number
        default: 12
        description: How many hours back to fetch (e.g. 12 = last 12 hours)
      - name: lake_id
        in: query
        type: string
        default: all
        description: Specific lake ID or "all" for every lake
      - name: format
        in: query
        type: string
        enum: [csv, xlsx]
        default: csv
    security:
      - Bearer: []
    responses:
      200:
        description: Downloadable data file
      204:
        description: No data found for the requested window
    """
    fmt      = request.args.get("format", "csv").lower()
    lake_id  = request.args.get("lake_id", "all").strip()
    try:
        hours = max(1, min(float(request.args.get("hours", 12)), 720))  # cap at 30 days
    except (ValueError, TypeError):
        hours = 12

    since = (datetime.now(tz=timezone.utc) - timedelta(hours=hours)).isoformat()

    query = {"timestamp": {"$gte": since}}
    if lake_id and lake_id.lower() != "all":
        query["lake_id"] = lake_id

    FIELDS = ["lake_id", "lake_name", "timestamp", "temperature",
              "rainfall", "water_level_rise", "velocity",
              "risk_score", "risk_level", "ml_score"]

    rows = list(
        db.telemetry.find(query, {"_id": 0})
        .sort("timestamp", 1)
        .limit(200_000)   # safety cap — ~200 k rows max
    )

    if not rows:
        return Response("No data found for the requested time window.",
                        mimetype="text/plain"), 204

    # ── Normalise rows — fill missing fields with empty string ────────────────
    clean_rows = []
    for r in rows:
        clean_rows.append({f: r.get(f, "") for f in FIELDS})

    # Filename stem
    lake_tag  = lake_id if lake_id.lower() != "all" else "all_lakes"
    ts_tag    = datetime.now(tz=timezone.utc).strftime("%Y%m%d_%H%M")
    stem      = f"glof_{lake_tag}_{int(hours)}h_{ts_tag}"

    # ── CSV ───────────────────────────────────────────────────────────────────
    if fmt == "csv":
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(clean_rows)
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={stem}.csv"},
        )

    # ── XLSX ──────────────────────────────────────────────────────────────────
    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return jsonify({"error": "openpyxl not installed on server"}), 500

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"GLOF Telemetry {int(hours)}h"

        # ── Header row styling ────────────────────────────────────────────────
        HEADER_FILL  = PatternFill("solid", fgColor="1A3C5E")
        HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        CELL_ALIGN   = Alignment(horizontal="center", vertical="center")
        THIN         = Side(style="thin", color="CBD5E1")
        BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        # Column display names
        COL_LABELS = {
            "lake_id":          "Lake ID",
            "lake_name":        "Lake Name",
            "timestamp":        "Timestamp (UTC)",
            "temperature":      "Temp (°C)",
            "rainfall":         "Rainfall (mm)",
            "water_level_rise": "Water Level Rise (cm)",
            "velocity":         "Velocity (cm/reading)",
            "risk_score":       "Risk Score",
            "risk_level":       "Risk Level",
            "ml_score":         "ML Score",
        }

        ws.append([COL_LABELS.get(f, f) for f in FIELDS])
        for col_idx, _ in enumerate(FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = CELL_ALIGN
            cell.border = BORDER

        # Freeze header row
        ws.freeze_panes = "A2"

        # ── Data rows ─────────────────────────────────────────────────────────
        RISK_COLORS = {
            "Low":      "D1FAE5",   # green tint
            "Moderate": "FEF3C7",   # yellow tint
            "High":     "FFEDD5",   # orange tint
            "Critical": "FEE2E2",   # red tint
            "Emergency":"FCE7F3",   # pink tint
        }

        for r_idx, row in enumerate(clean_rows, start=2):
            level = str(row.get("risk_level", ""))
            row_fill = PatternFill("solid", fgColor=RISK_COLORS.get(level, "FFFFFF")) if level else None
            for c_idx, field in enumerate(FIELDS, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row.get(field, ""))
                cell.border    = BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if row_fill:
                    cell.fill = row_fill

        # ── Auto-fit column widths (heuristic) ────────────────────────────────
        for col_idx, field in enumerate(FIELDS, start=1):
            max_len = max(
                len(COL_LABELS.get(field, field)),
                max((len(str(r.get(field, ""))) for r in clean_rows[:500]), default=0),
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        # ── Summary sheet ─────────────────────────────────────────────────────
        ws_sum = wb.create_sheet("Summary")
        ws_sum.append(["Export Info", ""])
        ws_sum.append(["Generated at (UTC)", datetime.now(tz=timezone.utc).isoformat()])
        ws_sum.append(["Hours window", hours])
        ws_sum.append(["Lake filter", lake_id])
        ws_sum.append(["Total rows", len(clean_rows)])
        ws_sum.append(["From (approx)", since])
        ws_sum.append(["Columns", ", ".join(FIELDS)])
        ws_sum.column_dimensions["A"].width = 24
        ws_sum.column_dimensions["B"].width = 50
        for cell in ws_sum["A"]:
            cell.font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={stem}.xlsx"},
        )

    return jsonify({"error": f"Unsupported format: {fmt}. Use csv or xlsx."}), 400


# ─── Alert Acknowledgement ────────────────────────────────────────────────────
@app.route("/api/alerts/acknowledge/<alert_id>", methods=["PATCH"])
@jwt_required()
def acknowledge_alert(alert_id):
    """
    Acknowledge an open alert (state: OPEN → ACKNOWLEDGED).
    ---
    tags:
      - Alerts
    parameters:
      - name: alert_id
        in: path
        type: string
        required: true
    security:
      - Bearer: []
    responses:
      200:
        description: Alert acknowledged
      404:
        description: Alert not found
    """
    try:
        obj_id = ObjectId(alert_id)
    except Exception:
        return jsonify({"error": "Invalid alert ID"}), 400

    user = get_jwt_identity()
    now = datetime.now(tz=timezone.utc).isoformat()

    result = db.alerts.update_one(
        {"_id": obj_id, "status": "OPEN"},
        {"$set": {"status": "ACKNOWLEDGED", "acknowledged_by": user, "acknowledged_at": now}},
    )
    if result.modified_count == 0:
        return jsonify({"error": "Alert not found or already acknowledged"}), 404

    write_audit_log("alert.acknowledged", user, {"alert_id": alert_id})
    return jsonify({"message": "Alert acknowledged", "acknowledged_by": user}), 200


# ─── Audit Log ────────────────────────────────────────────────────────────────
@app.route("/api/audit", methods=["GET"])
@jwt_required()
def get_audit_log():
    """
    Retrieve recent audit log entries (admin only).
    ---
    tags:
      - Admin
    security:
      - Bearer: []
    responses:
      200:
        description: List of audit entries
      403:
        description: Forbidden
    """
    claims = get_jwt()
    if claims.get("role") != "admin":
        return jsonify({"error": "Admin access required"}), 403

    limit = min(int(request.args.get("limit", 100)), 500)
    entries = list(
        db.audit_logs.find({}, {"_id": 0}).sort("timestamp", DESCENDING).limit(limit)
    )
    return jsonify(entries), 200


# ─── ML Model Status ──────────────────────────────────────────────────────────
@app.route("/api/ml/status", methods=["GET"])
@jwt_required()
def ml_status():
    """
    Returns the status and metadata of the loaded ML model.
    ---
    tags:
      - ML
    security:
      - Bearer: []
    responses:
      200:
        description: ML model info
    """
    from core.risk_engine import _ml_ready, _ml_features, _ml_classes
    import json as _json
    models_dir = os.path.join(os.path.dirname(__file__), "models")
    meta_path = os.path.join(models_dir, "model_meta.json")
    meta = {}
    if os.path.exists(meta_path):
        with open(meta_path) as f:
            meta = _json.load(f)
    return jsonify({
        "ml_ready": _ml_ready,
        "features": _ml_features,
        "classes": _ml_classes,
        "meta": meta,
        "blend_ratio": float(os.environ.get("ML_BLEND_RATIO", 0.4)),
    }), 200


# ─── Health Checks ────────────────────────────────────────────────────────────
@app.route("/health")
@limiter.exempt
def health():
    """
    Basic liveness check.
    ---
    tags:
      - System
    responses:
      200:
        description: Service is alive
    """
    return jsonify({
        "status": "ok",
        "version": "2.0.0",
        "time": datetime.now(tz=timezone.utc).isoformat(),
    }), 200


@app.route("/health/live")
@limiter.exempt
def health_live():
    """Kubernetes liveness probe."""
    return jsonify({"status": "alive"}), 200


@app.route("/health/ready")
@limiter.exempt
def health_ready():
    """
    Kubernetes readiness probe — checks DB and Redis.
    ---
    tags:
      - System
    responses:
      200:
        description: Service is ready
      503:
        description: Dependency not ready
    """
    checks = {"db": False, "redis": False}
    try:
        mongo_client.admin.command("ping")
        checks["db"] = True
    except Exception:
        pass
    try:
        redis_client.ping()
        checks["redis"] = True
    except Exception:
        pass

    ready = checks["db"]
    return jsonify({"status": "ready" if ready else "not_ready", "checks": checks}), 200 if ready else 503


@app.route("/health/detail")
@limiter.exempt
def health_detail():
    """
    Detailed health: DB + Redis status + last telemetry timestamp.
    ---
    tags:
      - System
    responses:
      200:
        description: Detailed health info
    """
    checks = {}
    try:
        mongo_client.admin.command("ping")
        checks["mongodb"] = "ok"
    except Exception as e:
        checks["mongodb"] = str(e)

    try:
        redis_client.ping()
        checks["redis"] = "ok"
    except Exception as e:
        checks["redis"] = str(e)

    last_tel = db.telemetry.find_one({}, {"_id": 0, "timestamp": 1}, sort=[("timestamp", DESCENDING)])
    checks["last_telemetry"] = last_tel["timestamp"] if last_tel else None

    return jsonify({
        "status": "ok" if all(v == "ok" for k, v in checks.items() if k != "last_telemetry") else "degraded",
        "version": "2.0.0",
        "checks": checks,
    }), 200


# ─── Time-Travel Snapshot (per-lake state N hours ago) ───────────────────────
@app.route("/api/telemetry/history-snapshot", methods=["GET"])
@jwt_required()
def telemetry_history_snapshot():
    """
    Return the most recent telemetry reading per lake from a point in the past.
    Used by the Time-Travel map slider on the frontend.
    ---
    tags:
      - Telemetry
    parameters:
      - name: hours_ago
        in: query
        type: number
        default: 0
        description: How many hours in the past (0 = now, 24 = 24h ago)
    security:
      - Bearer: []
    responses:
      200:
        description: List of per-lake snapshots
    """
    try:
        hours_ago = max(0.0, min(float(request.args.get("hours_ago", 0)), 72))
    except (ValueError, TypeError):
        hours_ago = 0.0

    now = datetime.now(tz=timezone.utc)
    # Target timestamp = now - hours_ago
    target_ts = now - timedelta(hours=hours_ago)
    # Look in a ±window around the target to find readings
    # Use a 30-min window; widen to 2h for older targets
    window_minutes = 30 if hours_ago < 6 else 120
    window_start = (target_ts - timedelta(minutes=window_minutes)).isoformat()
    window_end   = target_ts.isoformat()

    # For hours_ago=0 we just want the absolute latest per lake
    if hours_ago < 0.01:
        window_end = now.isoformat()
        window_start = (now - timedelta(minutes=30)).isoformat()

    pipeline = [
        {"$match": {"timestamp": {"$gte": window_start, "$lte": window_end}}},
        {"$sort": {"timestamp": -1}},
        {"$group": {
            "_id": "$lake_id",
            "lake_id":          {"$first": "$lake_id"},
            "lake_name":        {"$first": "$lake_name"},
            "timestamp":        {"$first": "$timestamp"},
            "risk_score":       {"$first": "$risk_score"},
            "risk_level":       {"$first": "$risk_level"},
            "water_level_rise": {"$first": "$water_level_rise"},
            "temperature":      {"$first": "$temperature"},
            "rainfall":         {"$first": "$rainfall"},
        }},
        {"$project": {"_id": 0}},
    ]

    results = list(db.telemetry.aggregate(pipeline))
    return jsonify({
        "hours_ago":  hours_ago,
        "target_ts":  target_ts.isoformat(),
        "window_min": window_start,
        "window_max": window_end,
        "snapshots":  results,
        "count":      len(results),
    }), 200


# ─── GLOF-Bot AI Chat ─────────────────────────────────────────────────────────
@app.route("/api/chat", methods=["POST"])
@jwt_required()
@limiter.limit("30 per minute")
def glof_chat():
    """
    GLOF-Bot: Gemini-powered AI assistant with live lake context.
    ---
    tags:
      - AI
    parameters:
      - in: body
        name: body
        required: true
        schema:
          properties:
            message: {type: string}
            history:
              type: array
              items:
                properties:
                  role: {type: string, enum: [user, model]}
                  text: {type: string}
    security:
      - Bearer: []
    responses:
      200:
        description: AI reply
      503:
        description: Gemini API not configured
    """
    body = request.get_json(silent=True) or {}
    user_message = (body.get("message") or "").strip()
    history      = body.get("history", []) or []

    if not user_message:
        return jsonify({"error": "message is required"}), 400

    gemini_key = os.environ.get("GEMINI_API_KEY", "")

    # ── Build lake context (top 12 lakes by risk score) ───────────────────────
    lakes_ctx = list(
        db.lakes.find({}, {"_id": 0, "id": 1, "name": 1, "state": 1,
                           "current_risk_level": 1, "current_risk_score": 1,
                           "elevation_m": 1, "area_ha": 1, "river_basin": 1})
        .sort("current_risk_score", DESCENDING)
        .limit(12)
    )
    lake_lines = []
    for l in lakes_ctx:
        lake_lines.append(
            f"  • {l.get('name','?')} ({l.get('id','?')}) — {l.get('state','?')}, "
            f"Risk: {l.get('current_risk_level','?')} ({l.get('current_risk_score',0):.1f}), "
            f"Basin: {l.get('river_basin','?')}, Elev: {l.get('elevation_m','?')}m"
        )
    lake_context = "\n".join(lake_lines) if lake_lines else "No lake data available."

    active_alerts = db.alerts.count_documents({"resolved": False})
    critical_count = db.lakes.count_documents({"current_risk_level": "Critical"})

    system_prompt = f"""You are GLOF-Bot, the AI assistant for GLOFWatch — an Early Warning System for 
Glacial Lake Outburst Floods (GLOFs) in the Himalayas (India). You help emergency responders, 
researchers, and local officials understand lake risk levels and take action.

CURRENT SYSTEM STATUS (live data):
- Total active alerts: {active_alerts}
- Lakes at Critical risk: {critical_count}

MONITORED LAKES (sorted by risk, highest first):
{lake_context}

INSTRUCTIONS:
- Answer questions about specific lakes, risk levels, and flood preparedness.
- If a lake is at Critical risk, always emphasize urgency and evacuation protocols.
- If asked about evacuation routes, recommend moving to higher ground and contacting NDMA/SDMA.
- Keep responses concise (2-4 sentences max) and actionable.
- Use simple language — some users may be in the field.
- You CANNOT send alerts or perform actions — direct users to the Admin panel for that.
- If a question is unrelated to GLOF/floods/disasters, politely redirect.
- Risk levels in order: Low → Moderate → High → Critical → Emergency.
"""

    # ── Rule-based fallback (no API key) ────────────────────────────────────
    if not gemini_key:
        # Simple rule engine as fallback
        msg_lower = user_message.lower()
        highest = lakes_ctx[0] if lakes_ctx else None
        if any(w in msg_lower for w in ["highest", "worst", "critical", "most dangerous"]):
            if highest:
                reply = (f"The highest-risk lake right now is **{highest['name']}** "
                         f"({highest['id']}) in {highest['state']} with risk level "
                         f"**{highest['current_risk_level']}** (score: {highest['current_risk_score']:.1f}). "
                         f"It is in the {highest['river_basin']} river basin.")
            else:
                reply = "No lake data is available at the moment."
        elif any(w in msg_lower for w in ["alert", "alerts", "active"]):
            reply = f"There are currently **{active_alerts} active unresolved alert(s)** in the system. You can view and acknowledge them in the Alerts panel."
        elif any(w in msg_lower for w in ["safe", "all safe", "ok"]):
            if critical_count == 0:
                reply = "✅ All monitored lakes are currently below Critical risk. Continue monitoring."
            else:
                reply = f"⚠️ **{critical_count} lake(s) are at Critical risk** right now. Immediate monitoring is advised."
        elif any(w in msg_lower for w in ["evacuate", "evacuation", "escape", "run"]):
            reply = ("In case of a GLOF emergency: **move to higher ground immediately**, avoid river banks and valleys. "
                     "Contact NDMA (011-26701700) or SDMA for your state. Do not wait for official confirmation if water levels are rising fast.")
        elif any(w in msg_lower for w in ["hello", "hi", "hey", "help"]):
            reply = ("Hello! I'm GLOF-Bot 🏔️. I can help you check lake risk levels, understand alerts, "
                     "and advise on flood preparedness. Try asking: *'Which lake has the highest risk?'*")
        elif any(w in msg_lower for w in ["how many lakes", "total lakes", "count"]):
            reply = f"GLOFWatch is currently monitoring **{len(lakes_ctx)}+ glacial lakes** across the Himalayan region."
        else:
            reply = ("I'm GLOF-Bot. I can answer questions about lake risk levels, active alerts, and flood preparedness. "
                     "Try: *'What is the current risk for GL001?'* or *'Which lake is most dangerous?'*")
        return jsonify({"reply": reply, "mode": "rule-based"}), 200

    # ── Gemini API call ───────────────────────────────────────────────────────
    try:
        import google.generativeai as genai
        genai.configure(api_key=gemini_key)
        model = genai.GenerativeModel(
            model_name="gemini-1.5-flash",
            system_instruction=system_prompt,
        )

        # Build chat history for Gemini
        gemini_history = []
        for h in history[-10:]:  # last 10 turns max
            role = "user" if h.get("role") == "user" else "model"
            gemini_history.append({"role": role, "parts": [h.get("text", "")]})

        chat = model.start_chat(history=gemini_history)
        response = chat.send_message(user_message)
        reply = response.text

        return jsonify({"reply": reply, "mode": "gemini"}), 200

    except Exception as e:
        app_log.warning(f"Gemini API error: {e}")
        return jsonify({"error": f"AI service error: {str(e)}", "reply": "I'm having trouble connecting to the AI service right now. Please try again in a moment."}), 503


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)

